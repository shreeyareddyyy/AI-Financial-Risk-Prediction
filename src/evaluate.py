"""
evaluate.py

Compares VADER's sentiment predictions (already in your report's
"Sentiment" column) against your own manually-labeled ground truth
(the "Manual Sentiment" column you filled in), and prints accuracy,
precision/recall/F1, and a confusion matrix - overall and per topic.

Expects an Excel file with at least these columns:
    Sentiment          - VADER's predicted label (Positive/Negative/Neutral)
    Manual Sentiment   - your hand-labeled ground truth (same 3 values)
    Topics             - topic tag(s), e.g. "Gold", "Fraud", "Crypto, Market"

Usage:
    python evaluate.py
    python evaluate.py --file labeled_sample.xlsx --sheet "Labeling Sample"
"""

import argparse
import sys

import openpyxl
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score

LABELS = ["Positive", "Neutral", "Negative"]


def load_rows(path, sheet_name, predicted_col):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [c.value for c in ws[1]]

    required = [predicted_col, "Manual Sentiment", "Topics"]
    for col in required:
        if col not in headers:
            print(f"ERROR: expected column '{col}' not found. "
                  f"Columns present: {headers}")
            sys.exit(1)

    sent_idx = headers.index(predicted_col)
    manual_idx = headers.index("Manual Sentiment")
    topic_idx = headers.index("Topics")

    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        predicted = row[sent_idx]
        manual = row[manual_idx]
        topic = row[topic_idx] or "Unknown"

        # skip rows you haven't labeled yet, or that are missing a
        # VADER prediction for some reason
        if not manual or not predicted:
            skipped += 1
            continue
        if manual not in LABELS or predicted not in LABELS:
            skipped += 1
            continue

        rows.append((predicted, manual, topic))

    if skipped:
        print(f"Note: skipped {skipped} rows (unlabeled, or label not in "
              f"{LABELS}). Fill in every row's 'Manual Sentiment' to "
              f"include it in the evaluation.\n")

    return rows


def print_report(y_true, y_pred, title):
    print(f"=== {title} (n={len(y_true)}) ===")
    if len(set(y_true)) < 2:
        print("Not enough label variety in this slice to score meaningfully.\n")
        return
    print("Accuracy:", round(accuracy_score(y_true, y_pred), 3))
    print()
    print(classification_report(y_true, y_pred, labels=LABELS, zero_division=0))
    print("Confusion matrix (rows = actual, cols = predicted)")
    print("Labels order:", LABELS)
    cm = confusion_matrix(y_true, y_pred, labels=LABELS)
    for label, counts in zip(LABELS, cm):
        print(f"  {label:<10} {counts}")
    print()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="labeled_sample.xlsx")
    parser.add_argument("--sheet", default="Labeling Sample")
    parser.add_argument("--column", default="Sentiment",
                         help="which column holds the model's predicted "
                              "label, e.g. 'Sentiment' (VADER) or "
                              "'Transformer Sentiment' (transformer)")
    args = parser.parse_args()

    rows = load_rows(args.file, args.sheet, args.column)
    if not rows:
        print("No labeled rows found - fill in 'Manual Sentiment' first.")
        return

    predicted = [r[0] for r in rows]
    manual = [r[1] for r in rows]
    topics = [r[2] for r in rows]

    # overall
    print_report(manual, predicted, f"OVERALL (predictions from '{args.column}')")

    # per topic (splits on comma since your Topics column can hold
    # combinations like "Fraud, Market" - each combination is treated as
    # its own slice here; that's fine for a diagnostic breakdown)
    by_topic = {}
    for p, m, t in zip(predicted, manual, topics):
        by_topic.setdefault(t, {"pred": [], "true": []})
        by_topic[t]["pred"].append(p)
        by_topic[t]["true"].append(m)

    for topic, data in sorted(by_topic.items(), key=lambda kv: -len(kv[1]["true"])):
        print_report(data["true"], data["pred"], f"TOPIC: {topic}")


if __name__ == "__main__":
    main()