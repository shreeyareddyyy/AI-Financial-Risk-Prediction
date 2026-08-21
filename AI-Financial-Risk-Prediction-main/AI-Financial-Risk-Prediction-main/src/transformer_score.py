"""
transformer_score.py

Scores your existing labeled_sample.xlsx with a pretrained transformer
sentiment model (cardiffnlp/twitter-roberta-base-sentiment-latest),
completely separately from your main pipeline - this does NOT touch
sentiment_pipeline.py or run_pipeline.py, so there's zero risk to what's
already working.

Why this model: it's a RoBERTa model fine-tuned specifically on ~124M
tweets for 3-class sentiment (negative/neutral/positive), so unlike raw
VADER it captures context, not just keyword matching - and it's small
enough to run reasonably on CPU for ~150 short texts.

What it does:
  1. Loads labeled_sample.xlsx (same file evaluate.py uses).
  2. Strips HTML tags from the Text column (your news items have
     raw <a href=...> tags mixed into the text from RSS - stripping
     them gives the model cleaner input than VADER got, which is
     worth mentioning as a methodology note, not hidden).
  3. Runs every row through the transformer in batches.
  4. Writes a new column "Transformer Sentiment" (Positive/Neutral/
     Negative, matching your Manual Sentiment format).
  5. Saves as labeled_sample_with_transformer.xlsx - a NEW file, your
     original labeled_sample.xlsx is untouched.

Usage:
    python transformer_score.py
    python transformer_score.py --file labeled_sample.xlsx --sheet "Labeling Sample"

After this finishes, run evaluate.py against the new file/column to get
the same accuracy/precision/recall/F1/confusion-matrix comparison you
already have for VADER - see the printed instructions at the end.
"""

import argparse
import re
import sys
import time

import openpyxl

LABEL_MAP = {
    "negative": "Negative",
    "neutral": "Neutral",
    "positive": "Positive",
}

HTML_TAG_RE = re.compile(r"<[^>]+>")


def clean_text(text: str) -> str:
    text = HTML_TAG_RE.sub(" ", str(text))
    text = text.replace("&nbsp;", " ").replace("&quot;", '"').replace("&amp;", "&")
    return " ".join(text.split()).strip()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="labeled_sample.xlsx")
    parser.add_argument("--sheet", default="Labeling Sample")
    parser.add_argument("--out", default=None,
                         help="output filename (default: overwrites --file "
                              "in place, adding/replacing this model's column)")
    parser.add_argument("--batch-size", type=int, default=16)
    parser.add_argument("--model", default="cardiffnlp/twitter-roberta-base-sentiment-latest",
                         help="HuggingFace model id to use for scoring")
    parser.add_argument("--column-name", default=None,
                         help="name for the output column (default: derived "
                              "from the model name)")
    args = parser.parse_args()

    if args.out is None:
        args.out = args.file  # overwrite in place by default so predictions accumulate
    if args.column_name is None:
        # turn e.g. "ProsusAI/finbert" into "FinBERT Sentiment"
        short_name = args.model.split("/")[-1]
        args.column_name = f"{short_name} Sentiment"

    print(f"Loading model: {args.model}")
    print("(first run of a given model downloads it, cached afterwards)...")
    try:
        from transformers import pipeline
    except ImportError:
        print("ERROR: transformers not installed. Run:\n"
              "  pip install transformers torch")
        sys.exit(1)

    t0 = time.time()
    classifier = pipeline(
        "sentiment-analysis",
        model=args.model,
        truncation=True,
        max_length=512,
    )
    print(f"Model loaded in {time.time() - t0:.1f}s\n")

    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers = [c.value for c in ws[1]]

    if "Text" not in headers:
        print(f"ERROR: 'Text' column not found. Columns present: {headers}")
        sys.exit(1)
    text_idx = headers.index("Text")

    rows = list(ws.iter_rows(min_row=2))
    texts = [clean_text(r[text_idx].value) for r in rows]
    print(f"Scoring {len(texts)} rows with column name '{args.column_name}'...")

    # batch through the classifier so it's not one HTTP-model-call per row
    predictions = []
    for i in range(0, len(texts), args.batch_size):
        batch = texts[i:i + args.batch_size]
        # skip empty strings, the pipeline errors on them
        safe_batch = [t if t else "." for t in batch]
        results = classifier(safe_batch)
        predictions.extend(LABEL_MAP.get(r["label"].lower(), "Neutral") for r in results)
        done = min(i + args.batch_size, len(texts))
        print(f"  {done}/{len(texts)} done")

    # write results into a column - reuse the column if it already exists
    # (e.g. re-running the same model), otherwise append a new one
    if args.column_name in headers:
        col_idx = headers.index(args.column_name) + 1
    else:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=args.column_name)

    for row_obj, pred in zip(rows, predictions):
        ws.cell(row=row_obj[0].row, column=col_idx, value=pred)

    wb.save(args.out)
    print(f"\nSaved: {args.out}  (column: '{args.column_name}')")
    print(f"\nCompare with: python evaluate.py --file {args.out} --column \"{args.column_name}\"")


if __name__ == "__main__":
    main()