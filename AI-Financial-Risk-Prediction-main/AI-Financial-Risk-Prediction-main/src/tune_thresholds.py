"""
tune_thresholds.py

VADER's raw compound Score (in your report's "Score" column) is a
continuous number roughly between -1 and +1. Your pipeline currently
converts that into Positive/Neutral/Negative using VADER's generic
default cutoffs (0.05 / -0.05). Those defaults were tuned on general
social-media text, not your specific gold/crypto/fraud/market dataset -
so they may not be the best cutoffs for YOUR data.

This script tries many different (pos_threshold, neg_threshold)
combinations against your own Score column, checks each one's accuracy
against your manually-labeled "Manual Sentiment" column, and reports
the combination that performs best - a standard, defensible technique
called threshold calibration.

Expects the same file/columns as evaluate.py:
    Score              - VADER's raw compound score (float)
    Manual Sentiment   - your hand-labeled ground truth

Usage:
    python tune_thresholds.py
    python tune_thresholds.py --file labeled_sample.xlsx --sheet "Labeling Sample"
"""

import argparse

import numpy as np
import openpyxl
from sklearn.metrics import accuracy_score, f1_score

LABELS = ["Positive", "Neutral", "Negative"]


def score_to_label(score, pos_thresh, neg_thresh):
    if score >= pos_thresh:
        return "Positive"
    elif score <= neg_thresh:
        return "Negative"
    return "Neutral"


def load_data(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [c.value for c in ws[1]]

    for col in ["Score", "Manual Sentiment"]:
        if col not in headers:
            raise SystemExit(f"ERROR: expected column '{col}' not found. "
                              f"Columns present: {headers}")

    score_idx = headers.index("Score")
    manual_idx = headers.index("Manual Sentiment")

    scores, manual = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        s, m = row[score_idx], row[manual_idx]
        if s is None or m not in LABELS:
            continue
        scores.append(float(s))
        manual.append(m)

    return np.array(scores), manual


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="labeled_sample.xlsx")
    parser.add_argument("--sheet", default="Labeling Sample")
    parser.add_argument("--step", type=float, default=0.02,
                         help="grid search step size (default 0.02)")
    args = parser.parse_args()

    scores, manual = load_data(args.file, args.sheet)
    print(f"Loaded {len(scores)} labeled rows with a Score value.\n")

    # baseline: VADER's own default thresholds
    baseline_pred = [score_to_label(s, 0.05, -0.05) for s in scores]
    baseline_acc = accuracy_score(manual, baseline_pred)
    baseline_f1 = f1_score(manual, baseline_pred, labels=LABELS, average="macro", zero_division=0)
    print(f"Baseline (VADER default thresholds 0.05 / -0.05):")
    print(f"  Accuracy = {baseline_acc:.3f}   Macro F1 = {baseline_f1:.3f}\n")

    # grid search: try pos_thresh from 0 to 0.6, neg_thresh from -0.6 to 0
    pos_range = np.arange(0.0, 0.61, args.step)
    neg_range = np.arange(-0.6, 0.01, args.step)

    best = {"acc": -1, "f1": -1, "pos": None, "neg": None}
    results = []

    for pos_t in pos_range:
        for neg_t in neg_range:
            if neg_t >= pos_t:
                continue  # nonsensical: negative cutoff must be below positive cutoff
            preds = [score_to_label(s, pos_t, neg_t) for s in scores]
            acc = accuracy_score(manual, preds)
            f1 = f1_score(manual, preds, labels=LABELS, average="macro", zero_division=0)
            results.append((acc, f1, pos_t, neg_t))
            if acc > best["acc"]:
                best = {"acc": acc, "f1": f1, "pos": pos_t, "neg": neg_t}

    print(f"Best thresholds found (by accuracy):")
    print(f"  pos_threshold = {best['pos']:.2f}   neg_threshold = {best['neg']:.2f}")
    print(f"  Accuracy = {best['acc']:.3f}   Macro F1 = {best['f1']:.3f}")
    print(f"  Improvement over baseline: {best['acc'] - baseline_acc:+.3f} accuracy, "
          f"{best['f1'] - baseline_f1:+.3f} macro F1\n")

    # also show the best-by-F1 in case it differs (F1 cares about balanced
    # performance across classes, accuracy can be skewed by a big class)
    best_f1_row = max(results, key=lambda r: r[1])
    if (best_f1_row[2], best_f1_row[3]) != (best["pos"], best["neg"]):
        print(f"Best thresholds found (by macro F1 instead):")
        print(f"  pos_threshold = {best_f1_row[2]:.2f}   neg_threshold = {best_f1_row[3]:.2f}")
        print(f"  Accuracy = {best_f1_row[0]:.3f}   Macro F1 = {best_f1_row[1]:.3f}\n")

    print("To apply: update score_to_label()'s thresholds in your pipeline "
          "to the values above, or apply them as a post-processing step "
          "when reading the Score column.")


if __name__ == "__main__":
    main()