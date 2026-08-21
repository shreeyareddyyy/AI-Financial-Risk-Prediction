"""
Sentiment, Risk Score & Multi-Source Detection Module
------------------------------------------------------
Part of: AI-Based Price Trend, Volatility & Risk Prediction System
         for Gold & Cryptocurrencies with Fraud Detection

Owner: P Shreeya (Sentiment, Risk Score & Dashboard)

WHAT THIS DOES
--------------
Takes content from four kinds of sources - plain text/news, images
(social media screenshots), audio clips, and video clips - pulls the
text out of each one, and runs it through a finance-tuned sentiment +
risk engine. Every item gets:
  1. A sentiment label (Positive / Negative / Neutral) with a score
  2. A topic tag (Gold / Crypto / Market / Fraud / General)
  3. A risk flag (Low / Medium / High) based on finance-specific
     keywords (e.g. "rug pull", "SEC probe", "delisted", "hack")

This is the piece that plugs into the unified risk-score formula and
the "Event-Driven Detection" feature on the roadmap slide (flagging
news/sentiment spikes that coincide with abnormal price volatility).

DESIGN NOTE ON "CONNECTING TO SOCIAL MEDIA"
--------------------------------------------
Live authenticated scraping of Twitter/X, Instagram, or Telegram is
NOT realistic to build, get approved, and demo in one day - those
platforms require paid API tiers or developer review that takes days
to weeks. What professors actually expect (and what your own roadmap
already calls "simulation mode") is that the pipeline can INGEST
content from those platforms once you have it - a screenshot of a
tweet, a downloaded video clip, a voice note, a news headline - and
process it the same way live-streamed data would be processed. That's
what this module does. Swapping in a live API later (e.g. Reddit's
free API, or a paid Twitter tier) only means writing one more adapter
function that feeds into `analyze_text()` - the rest of the pipeline
doesn't change.
"""

import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

# ---------------------------------------------------------------------
# 1. FINANCE-DOMAIN LEXICON
#    VADER alone doesn't know finance slang. We boost/override its
#    score using keyword categories built specifically for this
#    project's four topics: Gold, Crypto, Market, Fraud.
# ---------------------------------------------------------------------

TOPIC_KEYWORDS = {
    "Fraud": [
        "scam", "rug pull", "ponzi", "fraud", "fraudulent", "phishing",
        "hack", "hacked", "exploit", "ransom", "laundering", "fake",
        "counterfeit", "insider trading", "manipulation", "pump and dump",
    ],
    "Crypto": [
        "bitcoin", "btc", "ethereum", "eth", "crypto", "cryptocurrency",
        "altcoin", "token", "blockchain", "wallet", "exchange", "defi",
        "nft", "mining", "halving", "stablecoin",
    ],
    "Gold": [
        "gold", "bullion", "xau", "precious metal", "gold price",
        "gold reserve", "gold etf", "sovereign gold bond",
    ],
    "Market": [
        "stock market", "nifty", "sensex", "nasdaq", "dow jones",
        "s&p", "shares", "equity", "ipo", "bull market", "bear market",
        "recession", "inflation", "interest rate", "fed", "rbi",
    ],
}

# Words that push sentiment sharply negative / positive in a finance
# context even when VADER treats them as mild.
RISK_BOOST_WORDS = {
    # word: (sentiment_adjustment, risk_weight)
    "crash": (-0.6, 3), "plunge": (-0.5, 3), "collapse": (-0.6, 3),
    "scam": (-0.7, 3), "rug pull": (-0.8, 3), "hacked": (-0.6, 3),
    "delisted": (-0.5, 2), "banned": (-0.4, 2), "frozen": (-0.4, 2),
    "investigation": (-0.3, 2), "probe": (-0.3, 2), "lawsuit": (-0.3, 2),
    "surge": (0.5, 1), "rally": (0.5, 1), "all-time high": (0.6, 1),
    "record high": (0.6, 1), "breakout": (0.4, 1),
}

RISK_LEVELS = ["Low", "Medium", "High"]


@dataclass
class AnalysisResult:
    source_type: str            # text | image | audio | video | news
    source_ref: str             # filename, URL, or first 40 chars of text
    raw_text: str
    sentiment_label: str        # Positive | Negative | Neutral
    sentiment_score: float      # -1.0 to +1.0
    topics: list = field(default_factory=list)
    risk_level: str = "Low"
    risk_score: int = 0
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


# ---------------------------------------------------------------------
# 2. CORE TEXT SENTIMENT + RISK ENGINE
#    Every adapter (image/audio/video/news) funnels into this one
#    function. This is the single place the "TF-IDF + Logistic
#    Regression" classifier from the roadmap can later be swapped in
#    for VADER without touching anything upstream.
# ---------------------------------------------------------------------

_analyzer = SentimentIntensityAnalyzer()


def _detect_topics(text_lower: str) -> list:
    found = []
    for topic, words in TOPIC_KEYWORDS.items():
        if any(w in text_lower for w in words):
            found.append(topic)
    return found or ["General"]


def _risk_adjustment(text_lower: str):
    """Returns (score_adjustment, risk_weight) summed over matched keywords."""
    adj, weight = 0.0, 0
    for phrase, (score_adj, risk_w) in RISK_BOOST_WORDS.items():
        if phrase in text_lower:
            adj += score_adj
            weight += risk_w
    return adj, weight


def analyze_text(text: str, source_type: str = "text", source_ref: str = None) -> AnalysisResult:
    """Core function: run sentiment + topic + risk detection on any string."""
    text = (text or "").strip()
    source_ref = source_ref or (text[:40] + "..." if len(text) > 40 else text)

    if not text:
        return AnalysisResult(source_type, source_ref, "", "Neutral", 0.0, ["General"], "Low", 0)

    text_lower = text.lower()
    base = _analyzer.polarity_scores(text)["compound"]
    adj, risk_weight = _risk_adjustment(text_lower)
    final_score = max(-1.0, min(1.0, base + adj))

    if final_score >= 0.15:
        label = "Positive"
    elif final_score <= -0.15:
        label = "Negative"
    else:
        label = "Neutral"

    topics = _detect_topics(text_lower)

    # Risk level: negative sentiment + keyword weight + fraud topic all raise it
    risk_score = risk_weight
    if final_score <= -0.4:
        risk_score += 2
    elif final_score <= -0.15:
        risk_score += 1
    if "Fraud" in topics:
        risk_score += 2

    if risk_score >= 5:
        risk_level = "High"
    elif risk_score >= 2:
        risk_level = "Medium"
    else:
        risk_level = "Low"

    return AnalysisResult(
        source_type=source_type,
        source_ref=source_ref,
        raw_text=text,
        sentiment_label=label,
        sentiment_score=round(final_score, 3),
        topics=topics,
        risk_level=risk_level,
        risk_score=risk_score,
    )


# ---------------------------------------------------------------------
# 3. SOURCE ADAPTERS - image / audio / video / news
#    Each one just extracts text, then calls analyze_text().
# ---------------------------------------------------------------------

def analyze_image(path: str) -> AnalysisResult:
    """OCR a screenshot (e.g. a tweet, Instagram post, or news clipping)."""
    import pytesseract
    from PIL import Image

    text = pytesseract.image_to_string(Image.open(path))
    return analyze_text(text, source_type="image", source_ref=os.path.basename(path))


def analyze_audio(path: str) -> AnalysisResult:
    """Transcribe a voice note / audio clip, then analyze the transcript.
    Needs internet (uses the free Google Web Speech API via SpeechRecognition).
    For an offline alternative, swap in `openai-whisper` (see README)."""
    import speech_recognition as sr

    recognizer = sr.Recognizer()
    with sr.AudioFile(path) as source:
        audio_data = recognizer.record(source)
    try:
        text = recognizer.recognize_google(audio_data)
    except sr.UnknownValueError:
        text = ""
    except sr.RequestError as e:
        raise RuntimeError(f"Speech recognition service unavailable: {e}")

    return analyze_text(text, source_type="audio", source_ref=os.path.basename(path))


def analyze_video(path: str) -> AnalysisResult:
    """Extract the audio track from a video clip, then transcribe + analyze it."""
    from pydub import AudioSegment
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        wav_path = os.path.join(tmp, "extracted.wav")
        # pydub uses ffmpeg under the hood to pull audio from any video container
        audio = AudioSegment.from_file(path)
        audio.export(wav_path, format="wav")
        result = analyze_audio(wav_path)

    result.source_type = "video"
    result.source_ref = os.path.basename(path)
    return result


def analyze_news_feed(feed_urls: list) -> list:
    """Pull headlines from public financial news RSS feeds (no API key needed)
    and analyze each one. Needs internet access to reach the feed URLs."""
    import feedparser

    results = []
    for url in feed_urls:
        feed = feedparser.parse(url)
        for entry in feed.entries:
            headline = entry.get("title", "")
            summary = entry.get("summary", "")
            combined = f"{headline}. {summary}"
            results.append(analyze_text(combined, source_type="news", source_ref=headline[:60]))
    return results


def build_google_news_rss_url(query: str, region: str = "IN", lang: str = "en") -> str:
    """Build a Google News RSS search URL for any keyword or phrase - no API
    key needed, this is just Google News' public search-as-RSS feature."""
    import urllib.parse
    encoded = urllib.parse.quote(query)
    return f"https://news.google.com/rss/search?q={encoded}&hl={lang}-{region}&gl={region}&ceid={region}:{lang}"


def analyze_news_by_keyword(keywords: list, region: str = "IN", lang: str = "en",
                             only_relevant: bool = True) -> list:
    """The "browse the news and take whatever's market/gold/crypto/fraud
    related" flow, same shape as analyze_youtube_by_keyword():
    1. Build a Google News search feed for each keyword
    2. Analyze every headline that comes back
    3. By default, drop anything that doesn't tag to Gold/Crypto/Market/Fraud
       (i.e. keep only items where _detect_topics() found more than "General")

    No API key needed - Google News RSS search is public. Point this at
    keywords like "gold price", "bitcoin crash", "sensex today", "stock
    market fraud india" and it does its own discovery, exactly like the
    YouTube keyword search.
    """
    results = []
    for kw in keywords:
        url = build_google_news_rss_url(kw, region, lang)
        results += analyze_news_feed([url])

    if only_relevant:
        results = [r for r in results if r.topics != ["General"]]

    return results


# Suggested free/no-auth financial news RSS feeds you can plug straight in:
DEFAULT_NEWS_FEEDS = [
    "https://news.google.com/rss/search?q=gold+price+OR+cryptocurrency+OR+stock+market+fraud&hl=en-IN&gl=IN&ceid=IN:en",
    "https://finance.yahoo.com/news/rssindex",
]


def analyze_telegram_channel(channel_username: str, api_id: int, api_hash: str,
                              limit: int = 50) -> list:
    """Pull recent messages from a PUBLIC Telegram channel and analyze each one.

    This is a genuinely live social-media connection with a same-day setup:
    1. Log in at https://my.telegram.org with your own phone number
    2. Under "API Development Tools", register any app name -> you get
       api_id and api_hash INSTANTLY, no approval queue, no cost
    3. Point this at any public crypto/gold/market channel, e.g.
       "@CoinDeskGlobal", "@WSJmarkets", "@Reuters"

    Requires: pip install telethon
    """
    from telethon.sync import TelegramClient

    results = []
    with TelegramClient("sentiment_session", api_id, api_hash) as client:
        for message in client.iter_messages(channel_username, limit=limit):
            if message.text:
                results.append(analyze_text(
                    message.text,
                    source_type="telegram",
                    source_ref=f"{channel_username}#{message.id}",
                ))
    return results


def analyze_youtube_comments(video_id: str, api_key: str, max_results: int = 50) -> list:
    """Pull top-level comments from a YouTube video (e.g. a market news clip,
    a crypto influencer's upload) and analyze each one.

    Free setup, same day:
    1. Go to https://console.cloud.google.com -> create a project
    2. Enable "YouTube Data API v3" -> create an API key (instant)
    3. Free quota is 10,000 units/day - each comment-list call costs 1 unit,
       so this comfortably covers a student project's demo volume

    Requires: pip install google-api-python-client
    """
    from googleapiclient.discovery import build

    youtube = build("youtube", "v3", developerKey=api_key)
    request = youtube.commentThreads().list(
        part="snippet", videoId=video_id, maxResults=max_results, textFormat="plainText"
    )
    response = request.execute()

    results = []
    for item in response.get("items", []):
        comment = item["snippet"]["topLevelComment"]["snippet"]["textDisplay"]
        results.append(analyze_text(comment, source_type="youtube", source_ref=f"{video_id}#{item['id']}"))
    return results


def search_youtube_videos(query: str, api_key: str, max_results: int = 10,
                           order: str = "relevance", published_after: str = None) -> list:
    """Search YouTube for videos matching a keyword (e.g. "gold price crash",
    "crypto fraud", "sensex today") instead of needing specific video IDs -
    this is the "browse and find relevant content" step.

    Returns a list of dicts: {video_id, title, description, channel, published_at}

    Cost note: unlike comment/video lookups (1 unit each), search.list costs
    100 quota units per call regardless of maxResults. With the free 10,000
    unit/day quota that's about 100 searches/day - comfortably enough for a
    handful of keyword sweeps a day, but don't loop this in a tight loop.

    published_after: optional ISO 8601 timestamp (e.g. "2026-08-01T00:00:00Z")
    to restrict results to recent uploads only - useful for catching breaking
    news/event-driven spikes rather than old evergreen videos.
    """
    from googleapiclient.discovery import build

    youtube = build("youtube", "v3", developerKey=api_key)
    kwargs = dict(q=query, part="snippet", type="video", maxResults=max_results, order=order)
    if published_after:
        kwargs["publishedAfter"] = published_after
    response = youtube.search().list(**kwargs).execute()

    videos = []
    for item in response.get("items", []):
        snippet = item["snippet"]
        videos.append({
            "video_id": item["id"]["videoId"],
            "title": snippet["title"],
            "description": snippet["description"],
            "channel": snippet["channelTitle"],
            "published_at": snippet["publishedAt"],
        })
    return videos


def analyze_youtube_by_keyword(query: str, api_key: str, max_videos: int = 5,
                                include_comments: bool = True, comments_per_video: int = 20,
                                published_after: str = None) -> list:
    """The full "browse YouTube and analyze relevant content" flow:
    1. Search YouTube for videos matching `query`
    2. Analyze each video's title + description
    3. Optionally pull and analyze that video's top comments too

    This is what lets the pipeline discover content on its own instead of
    needing a hand-picked list of video IDs - point it at "gold price today",
    "bitcoin crash", "stock market fraud india", etc. and it finds the videos.

    Quota cost per call: 100 units (search) + up to max_videos units
    (video-level analysis is free from the search response itself) +
    up to max_videos comment-list calls (1 unit each) if include_comments=True.
    """
    results = []
    videos = search_youtube_videos(query, api_key, max_results=max_videos, published_after=published_after)

    for v in videos:
        combined = f"{v['title']}. {v['description']}"
        results.append(analyze_text(
            combined, source_type="youtube_video",
            source_ref=f"{v['video_id']} ({v['channel']})",
        ))
        if include_comments:
            try:
                results += analyze_youtube_comments(v["video_id"], api_key, max_results=comments_per_video)
            except Exception as e:
                # Comments can be disabled on a given video - don't let that kill the whole sweep
                print(f"  (comments unavailable for {v['video_id']}: {e})")

    return results


# ---------------------------------------------------------------------
# 4. AGGREGATION + REPORTING
# ---------------------------------------------------------------------

def aggregate_risk(results: list) -> dict:
    """Roll many AnalysisResults into a per-topic risk summary -
    this is the piece that feeds the project's unified risk-score formula."""
    summary = {}
    for r in results:
        for topic in r.topics:
            s = summary.setdefault(topic, {"count": 0, "avg_sentiment": 0.0,
                                            "high_risk_count": 0, "total_risk_score": 0})
            s["count"] += 1
            s["avg_sentiment"] += r.sentiment_score
            s["total_risk_score"] += r.risk_score
            if r.risk_level == "High":
                s["high_risk_count"] += 1

    for topic, s in summary.items():
        s["avg_sentiment"] = round(s["avg_sentiment"] / s["count"], 3)
    return summary


def export_to_excel(results: list, out_path: str):
    """Write results + per-topic summary to an Excel workbook with
    colour-coded risk levels (Low/Medium/High)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Sentiment Results"

    headers = ["Timestamp", "Source Type", "Source", "Sentiment", "Score",
               "Topics", "Risk Level", "Risk Score", "Text"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fills = {
        "High": PatternFill(start_color="F8CBCB", end_color="F8CBCB", fill_type="solid"),
        "Medium": PatternFill(start_color="FCEBB6", end_color="FCEBB6", fill_type="solid"),
        "Low": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    }

    for r in results:
        row = [r.timestamp, r.source_type, r.source_ref, r.sentiment_label,
               r.sentiment_score, ", ".join(r.topics), r.risk_level,
               r.risk_score, r.raw_text[:500]]
        ws.append(row)
        ws.cell(row=ws.max_row, column=7).fill = fills.get(r.risk_level)

    for col, width in zip("ABCDEFGHI", [20, 12, 25, 12, 8, 20, 10, 10, 60]):
        ws.column_dimensions[col].width = width

    # Second sheet: per-topic summary
    summary = aggregate_risk(results)
    ws2 = wb.create_sheet("Topic Summary")
    ws2.append(["Topic", "Item Count", "Avg Sentiment", "High-Risk Items", "Total Risk Score"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for topic, s in summary.items():
        ws2.append([topic, s["count"], s["avg_sentiment"], s["high_risk_count"], s["total_risk_score"]])
    for col, width in zip("ABCDE", [15, 12, 14, 16, 16]):
        ws2.column_dimensions[col].width = width

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------
# 5. DEMO / SELF-TEST
# ---------------------------------------------------------------------

if __name__ == "__main__":
    sample_items = [
        "Bitcoin crashes 18% after major exchange hack, investigators suspect insider fraud.",
        "Gold prices hit an all-time high as investors seek safe haven amid market volatility.",
        "Sensex rallies 500 points as inflation data comes in better than expected.",
        "SEC opens investigation into crypto exchange over alleged rug pull scheme.",
        "Just had a great cup of coffee this morning, nothing special today.",
    ]

    results = [analyze_text(t) for t in sample_items]

    print(f"{'Sentiment':<10} {'Score':<7} {'Topics':<20} {'Risk':<8} Text")
    print("-" * 100)
    for r in results:
        print(f"{r.sentiment_label:<10} {r.sentiment_score:<7} {', '.join(r.topics):<20} {r.risk_level:<8} {r.raw_text[:60]}")

    out = export_to_excel(results, "demo_sentiment_report.xlsx")
    print(f"\nExcel report written to: {out}")